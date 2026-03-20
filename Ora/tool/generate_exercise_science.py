#!/usr/bin/env python3
import argparse
import json
import os
import re
import time
import urllib.error
import urllib.request

from openpyxl import load_workbook

# Paths relative to the project root
CATALOG_PATH = 'lib/data/seed/exercise_catalog_seed.json'
SCIENCE_PATH = 'lib/data/seed/exercise_science_seed.json'
ENV_PATH = '.env'

# Configuration
DEFAULT_MODEL = 'gpt-4o-mini'


def load_env(path):
    """Simple manual .env loader to avoid dependencies."""
    if not os.path.exists(path):
        return
    with open(path, 'r', encoding='utf-8') as handle:
        for line in handle:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if '=' in line:
                key, value = line.split('=', 1)
                os.environ[key.strip()] = value.strip().strip('"').strip("'")


def normalize_name_key(value):
    text = str(value or '').strip().lower()
    if not text:
        return ''
    return re.sub(r'[^a-z0-9]+', ' ', text).strip()


def build_catalog_lookup(catalog):
    lookup = {}
    for item in catalog:
        for candidate in [item.get('canonical_name'), *(item.get('aliases') or [])]:
            key = normalize_name_key(candidate)
            if key and key not in lookup:
                lookup[key] = item['canonical_name']
    return lookup


def normalize_string_list(value):
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    if value is None:
        return []
    text = str(value).strip()
    return [text] if text else []


def normalize_source_documents(value, fallback_citations):
    documents = []
    if isinstance(value, list):
        for item in value:
            if isinstance(item, str):
                citation = item.strip()
                if not citation:
                    continue
                documents.append({
                    'id': f'S{len(documents) + 1}',
                    'citation': citation,
                })
                continue
            if not isinstance(item, dict):
                text = str(item).strip()
                if not text:
                    continue
                documents.append({
                    'id': f'S{len(documents) + 1}',
                    'citation': text,
                })
                continue

            citation = str(
                item.get('citation')
                or item.get('reference')
                or item.get('text')
                or item.get('source')
                or ''
            ).strip()
            title = str(item.get('title') or '').strip()
            if not citation and title:
                citation = title
            if not citation:
                continue

            raw_id = str(
                item.get('id')
                or item.get('source_id')
                or item.get('document_id')
                or f'S{len(documents) + 1}'
            ).strip().upper().replace(' ', '')
            doc = {
                'id': raw_id or f'S{len(documents) + 1}',
                'citation': citation,
            }
            if title:
                doc['title'] = title
            document_type = str(item.get('document_type') or item.get('type') or '').strip()
            if document_type:
                doc['document_type'] = document_type
            url = str(item.get('url') or item.get('link') or '').strip()
            if url:
                doc['url'] = url
            year = item.get('year')
            if isinstance(year, int):
                doc['year'] = year
            elif isinstance(year, str) and year.strip().isdigit():
                doc['year'] = int(year.strip())
            relevance = str(
                item.get('relevance')
                or item.get('note')
                or item.get('summary')
                or ''
            ).strip()
            if relevance:
                doc['relevance'] = relevance
            documents.append(doc)

    if not documents:
        for citation in fallback_citations:
            documents.append({
                'id': f'S{len(documents) + 1}',
                'citation': citation,
            })

    normalized = []
    used_ids = set()
    for item in documents:
        doc_id = str(item.get('id') or f'S{len(normalized) + 1}').strip().upper().replace(' ', '')
        if not doc_id or doc_id in used_ids:
            doc_id = f'S{len(normalized) + 1}'
        used_ids.add(doc_id)
        item['id'] = doc_id
        normalized.append(item)
    return normalized


def normalize_information_sections(value):
    if not isinstance(value, list):
        return []

    sections = []
    for raw_section in value:
        if not isinstance(raw_section, dict):
            continue
        raw_id = str(raw_section.get('id') or raw_section.get('key') or raw_section.get('slug') or '').strip()
        title = str(raw_section.get('title') or raw_section.get('label') or '').strip()
        if not raw_id and not title:
            continue
        section_id = re.sub(r'[^a-z0-9]+', '_', (raw_id or title).strip().lower()).strip('_')
        if not section_id:
            continue
        section = {
            'id': section_id,
            'title': title or section_id.replace('_', ' ').title(),
            'items': [],
        }
        summary = str(
            raw_section.get('summary')
            or raw_section.get('overview')
            or raw_section.get('description')
            or ''
        ).strip()
        if summary:
            section['summary'] = summary

        raw_items = raw_section.get('items') or raw_section.get('claims') or raw_section.get('points') or []
        if isinstance(raw_items, list):
            for raw_item in raw_items:
                if isinstance(raw_item, str):
                    text = raw_item.strip()
                    if not text:
                        continue
                    section['items'].append({'title': text, 'source_ids': []})
                    continue
                if not isinstance(raw_item, dict):
                    text = str(raw_item).strip()
                    if not text:
                        continue
                    section['items'].append({'title': text, 'source_ids': []})
                    continue

                title_text = str(
                    raw_item.get('title')
                    or raw_item.get('claim')
                    or raw_item.get('text')
                    or raw_item.get('label')
                    or raw_item.get('summary')
                    or ''
                ).strip()
                detail_text = str(
                    raw_item.get('detail')
                    or raw_item.get('body')
                    or raw_item.get('note')
                    or raw_item.get('description')
                    or raw_item.get('explanation')
                    or ''
                ).strip()
                if not title_text and not detail_text:
                    continue
                point = {
                    'title': title_text or detail_text,
                    'source_ids': [],
                }
                if title_text and detail_text:
                    point['detail'] = detail_text
                source_ids = raw_item.get('source_ids') or raw_item.get('sources') or raw_item.get('document_ids') or []
                if isinstance(source_ids, list):
                    point['source_ids'] = [
                        str(item).strip().upper().replace(' ', '')
                        for item in source_ids
                        if str(item).strip()
                    ]
                elif source_ids:
                    point['source_ids'] = [str(source_ids).strip().upper().replace(' ', '')]
                section['items'].append(point)

        sections.append(section)
    return sections


def repair_and_validate_json(raw_text, exercise_name):
    """Attempts to extract, repair, and validate the JSON response from the LLM."""
    json_match = re.search(r'```json\s*(.*?)\s*```', raw_text, re.DOTALL)
    if json_match:
        raw_text = json_match.group(1)
    else:
        brace_match = re.search(r'(\{.*\})', raw_text, re.DOTALL)
        if brace_match:
            raw_text = brace_match.group(1)

    try:
        data = json.loads(raw_text.strip())
    except json.JSONDecodeError as exc:
        print(f'  -> JSON Syntax Error: {exc}. Attempting basic character repair...')
        repaired = re.sub(r',\s*([\]}])', r'\1', raw_text.strip())
        try:
            data = json.loads(repaired)
        except json.JSONDecodeError:
            return None

    if not isinstance(data, dict):
        return None

    data['canonical_name'] = str(data.get('canonical_name') or exercise_name)
    data['instructions'] = normalize_string_list(data.get('instructions'))
    data['avoid'] = normalize_string_list(data.get('avoid'))
    data['citations'] = normalize_string_list(data.get('citations'))
    data['visual_asset_paths'] = normalize_string_list(data.get('visual_asset_paths'))
    data['information_sections'] = normalize_information_sections(
        data.get('information_sections') or data.get('sections')
    )
    data['source_documents'] = normalize_source_documents(
        data.get('source_documents'),
        data['citations'],
    )
    return data


def extract_exercises_from_xlsx(path):
    workbook = load_workbook(path, data_only=False)
    sheet = workbook[workbook.sheetnames[0]]
    names = []
    seen = set()
    for row in sheet.iter_rows(values_only=True):
        first = row[0] if len(row) > 0 else None
        sets_value = row[2] if len(row) > 2 else None
        if first is None:
            continue
        name = str(first).strip()
        if not name:
            continue
        lowered = name.lower()
        if lowered.startswith('day '):
            continue
        if lowered in {
            'exercises',
            'repeat',
            'progression type',
            'weekly direct volume',
            'upper',
            'lower',
        }:
            continue
        if not isinstance(sets_value, (int, float)) or sets_value < 1:
            continue
        normalized = normalize_name_key(name)
        if normalized in seen:
            continue
        seen.add(normalized)
        names.append(name)
    return names


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('--xlsx-path', help='Optional workbook path to extract exercise names from.')
    parser.add_argument('--only-xlsx', action='store_true', help='Only generate for the exercises extracted from --xlsx-path.')
    return parser.parse_args()


def main():
    args = parse_args()
    load_env(ENV_PATH)
    api_key = os.environ.get('OPENAI_API_KEY')
    model = os.environ.get('OPENAI_MODEL', DEFAULT_MODEL)

    if not api_key:
        print('Error: OPENAI_API_KEY environment variable not set.')
        print("Usage: OPENAI_API_KEY='your_key' [OPENAI_MODEL='gpt-4o-mini'] python3 tool/generate_exercise_science.py [--xlsx-path workbook.xlsx --only-xlsx]")
        return

    if not os.path.exists(CATALOG_PATH):
        print(f'Error: Catalog file not found at {CATALOG_PATH}.')
        return

    with open(CATALOG_PATH, 'r', encoding='utf-8') as handle:
        catalog = json.load(handle)

    all_exercise_names = [item['canonical_name'] for item in catalog]
    catalog_lookup = build_catalog_lookup(catalog)
    print(f'Found {len(all_exercise_names)} exercises in catalog.')

    workbook_names = []
    if args.xlsx_path:
        workbook_path = args.xlsx_path
        if not os.path.exists(workbook_path):
            print(f'Error: Workbook file not found at {workbook_path}.')
            return
        workbook_names = extract_exercises_from_xlsx(workbook_path)
        print(f'Workbook exercises extracted: {len(workbook_names)}')
        missing_from_catalog = [
            name for name in workbook_names if normalize_name_key(name) not in catalog_lookup
        ]
        if missing_from_catalog:
            print('Workbook exercises missing from catalog:')
            for name in missing_from_catalog:
                print(f'  - {name}')

    existing_science = []
    if os.path.exists(SCIENCE_PATH):
        with open(SCIENCE_PATH, 'r', encoding='utf-8') as handle:
            try:
                existing_science = json.load(handle)
            except json.JSONDecodeError:
                existing_science = []

    completed_names = {}
    for item in existing_science:
        key = normalize_name_key(item.get('canonical_name') or item.get('exercise_name'))
        if key:
            completed_names[key] = item

    if args.only_xlsx:
        missing_names = [name for name in workbook_names if normalize_name_key(name) not in completed_names]
    else:
        missing_names = [name for name in all_exercise_names if normalize_name_key(name) not in completed_names]
        if workbook_names:
            workbook_missing = [
                name for name in workbook_names if normalize_name_key(name) not in completed_names
            ]
            merged = []
            seen = set()
            for name in workbook_missing + missing_names:
                key = normalize_name_key(name)
                if key in seen:
                    continue
                seen.add(key)
                merged.append(name)
            missing_names = merged

    print(f'Completed: {len(completed_names)} | Missing target set: {len(missing_names)}')

    if not missing_names:
        print('No missing exercises. Done!')
        return

    url = 'https://api.openai.com/v1/chat/completions'
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {api_key}',
    }

    try:
        for index, name in enumerate(missing_names, start=1):
            print(f'Processing ({index}/{len(missing_names)}): {name}...')

            payload = {
                'model': model,
                'messages': [
                    {
                        'role': 'system',
                        'content': 'You are a biomechanics research assistant. Output only valid JSON. Make conservative claims. Use empty lists instead of inventing unsupported evidence.',
                    },
                    {
                        'role': 'user',
                        'content': f"""Provide structured exercise-information data for '{name}'.
Return ONLY this JSON structure:
{{
  "canonical_name": "{name}",
  "instructions": ["Step 1", "Step 2", "Step 3"],
  "avoid": ["Mistake 1", "Mistake 2"],
  "information_sections": [
    {{
      "id": "safety",
      "title": "Safety",
      "summary": "One conservative paragraph.",
      "items": [
        {{
          "title": "Claim heading",
          "detail": "Short evidence-guided explanation.",
          "source_ids": ["S1"]
        }}
      ]
    }},
    {{
      "id": "effectiveness",
      "title": "Effectiveness",
      "summary": "One conservative paragraph.",
      "items": [
        {{
          "title": "Claim heading",
          "detail": "Short evidence-guided explanation.",
          "source_ids": ["S1"]
        }}
      ]
    }}
  ],
  "source_documents": [
    {{
      "id": "S1",
      "title": "Paper title",
      "citation": "Author (Year). Title. Journal.",
      "document_type": "Journal article",
      "year": 2020,
      "url": "",
      "relevance": "Why this document matters for critique."
    }}
  ],
  "citations": ["Author (Year). Title. Journal."],
  "visual_asset_paths": []
}}
Rules:
- Keep every claim conservative and tied to the cited document when possible.
- Prefer direct, critique-friendly phrasing over hype.
- If you lack support for a section, return an empty items list or omit the section instead of guessing.
- Keep source IDs aligned with the source_documents array.
""",
                    },
                ],
                'response_format': {'type': 'json_object'},
                'temperature': 0.2,
            }

            request = urllib.request.Request(
                url,
                data=json.dumps(payload).encode('utf-8'),
                headers=headers,
                method='POST',
            )

            try:
                with urllib.request.urlopen(request) as response:
                    response_json = json.loads(response.read().decode('utf-8'))
                    raw_content = response_json['choices'][0]['message']['content']
                    validated_data = repair_and_validate_json(raw_content, name)

                    if validated_data:
                        existing_science.append(validated_data)
                        with open(SCIENCE_PATH, 'w', encoding='utf-8') as handle:
                            json.dump(existing_science, handle, indent=2)
                        print('  -> Success.')
                    else:
                        print(f'  -> Failed validation/repair for: {name}')

            except urllib.error.HTTPError as exc:
                print(f'  -> HTTP Error: {exc.code}')
                try:
                    error_body = exc.read().decode('utf-8')
                    print(f'  -> Details: {error_body}')
                except Exception:
                    pass
                if exc.code == 429:
                    print('  -> Rate limited. Sleeping 20s...')
                    time.sleep(20)
                continue

            time.sleep(1)

    except KeyboardInterrupt:
        print('\nStopped. Progress saved.')

    print('\nDone!')


if __name__ == '__main__':
    main()
